"""
Import Tier 3 (Vitamins/Minerals) and Tier 4 (Processing Aids) into ingredients_dict.

UPSERT by canonical_name:
- If entry exists: UPDATE content fields + severity (preserve cluster_id, allergen_group, flags)
- If new: INSERT full entry

Severity mapping: Beneficial → good, Neutral → neutral, Caution → caution, Danger → danger
definition → base_description (confirmed same concept)

Usage:
    python3 scripts/import/import_tiers.py --dry-run   # preview changes
    python3 scripts/import/import_tiers.py              # execute changes
"""

import sys, os, json, re
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config import get_client

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
ERROR_LOG_PATH = SCRIPT_DIR / "import_errors.json"

SEVERITY_MAP = {
    'Beneficial': 'good',
    'beneficial': 'good',
    'Neutral': 'neutral',
    'neutral': 'neutral',
    'Caution': 'caution',
    'caution': 'caution',
    'Danger': 'danger',
    'danger': 'danger',
}

DRY_RUN = '--dry-run' in sys.argv


def split_citations(raw: str) -> list[str]:
    """Split semicolon-delimited citations string into a list."""
    if not raw:
        return []
    return [c.strip() for c in raw.split(';') if c.strip()]


def parse_tier3_xlsx(filepath):
    """Parse Tier 3 spreadsheet (69 entries)."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, canonical, display, dog_sev, cat_sev, pos_red, base_desc, dog_ctx, cat_ctx, pos_ctx, definition, tldr, detail_body, citations, concern = row
        if not canonical:
            continue
        entries.append({
            'canonical_name': canonical.strip(),
            'display_name': display,
            'dog_base_severity': SEVERITY_MAP.get(str(dog_sev).strip() if dog_sev else 'neutral', 'neutral'),
            'cat_base_severity': SEVERITY_MAP.get(str(cat_sev).strip() if cat_sev else 'neutral', 'neutral'),
            'position_reduction_eligible': bool(pos_red) if pos_red else False,
            'base_description': base_desc or definition,
            'dog_context': dog_ctx,
            'cat_context': cat_ctx,
            'position_context': pos_ctx,
            'definition': definition,
            'tldr': tldr,
            'detail_body': detail_body,
            'citations_display': split_citations(citations) if citations else [],
            'citation_sources': split_citations(citations) if citations else [],
            'tier': 'tier3',
        })
    return entries


def parse_tier4_markdown(filepath):
    """Parse Tier 4 markdown with code blocks (19 entries)."""
    with open(filepath, 'r') as f:
        content = f.read()

    # Extract code blocks
    blocks = re.findall(r'```\n(.*?)```', content, re.DOTALL)
    entries = []

    for block in blocks:
        entry = {}
        lines = block.strip().split('\n')
        current_key = None
        current_val = []

        for line in lines:
            match = re.match(r'^(\w[\w_]*)\s*:\s*(.*)$', line)
            if match:
                if current_key:
                    entry[current_key] = '\n'.join(current_val).strip()
                current_key = match.group(1)
                current_val = [match.group(2)]
            else:
                if current_key:
                    current_val.append(line)

        if current_key:
            entry[current_key] = '\n'.join(current_val).strip()

        if 'canonical_name' not in entry:
            continue

        raw_citations = entry.get('citations_display', '')

        entries.append({
            'canonical_name': entry.get('canonical_name', '').strip(),
            'display_name': entry.get('display_name', ''),
            'dog_base_severity': SEVERITY_MAP.get(entry.get('dog_base_severity', 'neutral'), 'neutral'),
            'cat_base_severity': SEVERITY_MAP.get(entry.get('cat_base_severity', 'neutral'), 'neutral'),
            'position_reduction_eligible': entry.get('position_reduction_eligible', 'FALSE').upper() == 'TRUE',
            'base_description': entry.get('base_description', '') or entry.get('definition', ''),
            'dog_context': entry.get('dog_context', ''),
            'cat_context': entry.get('cat_context', ''),
            'position_context': entry.get('position_context', ''),
            'definition': entry.get('definition', ''),
            'tldr': entry.get('tldr', ''),
            'detail_body': entry.get('detail_body', ''),
            'citations_display': split_citations(raw_citations),
            'citation_sources': split_citations(raw_citations),
            'tier': 'tier4',
        })

    return entries


def import_entries(sb, entries):
    """UPSERT entries into ingredients_dict."""
    stats = {
        'updated': 0,
        'inserted': 0,
        'severity_changed': 0,
        'skipped': 0,
        'errors': 0,
    }
    changes = []
    errors = []

    for entry in entries:
        cn = entry['canonical_name']

        # Check if exists
        result = sb.table('ingredients_dict') \
            .select('id, canonical_name, dog_base_severity, cat_base_severity, display_name, tldr') \
            .eq('canonical_name', cn) \
            .execute()

        # Content fields to set (never overwrites cluster_id, allergen_group, boolean flags)
        update_data = {
            'display_name': entry['display_name'],
            'dog_base_severity': entry['dog_base_severity'],
            'cat_base_severity': entry['cat_base_severity'],
            'position_reduction_eligible': entry['position_reduction_eligible'],
            'base_description': entry['base_description'],
            'dog_context': entry['dog_context'],
            'cat_context': entry['cat_context'],
            'definition': entry['definition'] or entry['base_description'],
            'tldr': entry['tldr'],
            'detail_body': entry['detail_body'],
            'citations_display': entry['citations_display'],  # JSONB array
            'citation_sources': entry['citation_sources'],     # TEXT[] array
            'position_context': entry['position_context'],
        }

        if result.data and len(result.data) > 0:
            existing = result.data[0]
            old_dog = existing.get('dog_base_severity')
            old_cat = existing.get('cat_base_severity')
            had_tldr = bool(existing.get('tldr'))

            severity_changed = (old_dog != entry['dog_base_severity'] or old_cat != entry['cat_base_severity'])

            if DRY_RUN:
                action = 'UPDATE'
                if severity_changed:
                    action += f' (severity: {old_dog}/{old_cat} → {entry["dog_base_severity"]}/{entry["cat_base_severity"]})'
                if not had_tldr:
                    action += ' (adding content)'
                print(f'  {cn}: {action}')
            else:
                try:
                    sb.table('ingredients_dict') \
                        .update(update_data) \
                        .eq('canonical_name', cn) \
                        .execute()
                except Exception as e:
                    print(f'  ERROR updating {cn}: {e}')
                    errors.append({'canonical_name': cn, 'action': 'update', 'error': str(e)})
                    stats['errors'] += 1
                    continue

            stats['updated'] += 1
            if severity_changed:
                stats['severity_changed'] += 1
                changes.append(f'{cn}: {old_dog}/{old_cat} → {entry["dog_base_severity"]}/{entry["cat_base_severity"]}')
        else:
            # INSERT new entry
            insert_data = {
                **update_data,
                'canonical_name': cn,
                'is_unnamed_species': False,
                'is_legume': False,
                'cat_carb_flag': False,
            }

            if DRY_RUN:
                print(f'  {cn}: INSERT (new) [{entry["dog_base_severity"]}/{entry["cat_base_severity"]}]')
            else:
                try:
                    sb.table('ingredients_dict') \
                        .insert(insert_data) \
                        .execute()
                except Exception as e:
                    print(f'  ERROR inserting {cn}: {e}')
                    errors.append({'canonical_name': cn, 'action': 'insert', 'error': str(e)})
                    stats['errors'] += 1
                    continue

            stats['inserted'] += 1

    return stats, changes, errors


def main():
    mode = 'DRY RUN' if DRY_RUN else 'LIVE'
    print(f'=== Tier 3 + Tier 4 Import ({mode}) ===\n')

    tier3_path = SCRIPT_DIR / 'TIER3_VITAMINS_MINERALS_UPDATED.xlsx'
    tier4_path = SCRIPT_DIR / 'TIER4_PROCESSING_AIDS.md'

    print('Parsing Tier 3 (xlsx)...')
    tier3 = parse_tier3_xlsx(tier3_path)
    print(f'  {len(tier3)} entries')

    print('Parsing Tier 4 (markdown)...')
    tier4 = parse_tier4_markdown(tier4_path)
    print(f'  {len(tier4)} entries')

    all_entries = tier3 + tier4
    print(f'\nTotal: {len(all_entries)} entries to import\n')

    sb = get_client()

    print('Importing Tier 3...')
    stats3, changes3, errors3 = import_entries(sb, tier3)

    print('\nImporting Tier 4...')
    stats4, changes4, errors4 = import_entries(sb, tier4)

    # Totals
    total_updated = stats3['updated'] + stats4['updated']
    total_inserted = stats3['inserted'] + stats4['inserted']
    total_sev = stats3['severity_changed'] + stats4['severity_changed']
    total_errors = stats3['errors'] + stats4['errors']
    all_changes = changes3 + changes4
    all_errors = errors3 + errors4

    print(f'\n{"="*60}')
    print(f'IMPORT REPORT ({mode})')
    print(f'{"="*60}')
    print(f'Tier 3: {stats3["updated"]} updated, {stats3["inserted"]} inserted, {stats3["severity_changed"]} severity changes')
    print(f'Tier 4: {stats4["updated"]} updated, {stats4["inserted"]} inserted, {stats4["severity_changed"]} severity changes')
    print(f'Total:  {total_updated} updated, {total_inserted} inserted, {total_sev} severity changes, {total_errors} errors')

    if all_changes:
        print(f'\nSeverity changes:')
        for c in all_changes:
            print(f'  {c}')

    if all_errors:
        with open(ERROR_LOG_PATH, 'w') as f:
            json.dump(all_errors, f, indent=2)
        print(f'\nErrors saved to {ERROR_LOG_PATH}')

    print(f'\nNote: brewer_dried_yeast is an alias for brewers_yeast — add to synonyms.json.')


if __name__ == '__main__':
    main()
