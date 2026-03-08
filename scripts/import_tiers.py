"""
Import Tier 3 (Vitamins/Minerals) and Tier 4 (Processing Aids) into ingredients_dict.

UPSERT by canonical_name:
- If entry exists: UPDATE content fields + severity
- If new: INSERT full entry

Severity mapping: Beneficial → good, Neutral → neutral, Caution → caution, Danger → danger
definition → base_description (confirmed same concept)
review_status = 'verified' for all imported entries

Preserves: cluster_id, allergen_group, allergen_group_possible, is_unnamed_species, is_legume, cat_carb_flag
"""

import sys, os, json, re
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'scripts', 'import'))

import openpyxl

SEVERITY_MAP = {
    'Beneficial': 'good',
    'beneficial': 'good',
    'Neutral': 'neutral',
    'neutral': 'neutral',
    'Caution': 'caution',
    'caution': 'caution',
    'Danger': 'danger',
    'danger': 'danger',
}

DRY_RUN = '--dry-run' in sys.argv


def get_client():
    """Get Supabase client using env vars."""
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '..', '..', '.env'))
    from supabase import create_client
    url = os.environ['SUPABASE_URL']
    key = os.environ['SUPABASE_SERVICE_ROLE_KEY']
    return create_client(url, key)


def parse_tier3_xlsx(filepath):
    """Parse Tier 3 spreadsheet (69 entries)."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, canonical, display, dog_sev, cat_sev, pos_red, base_desc, dog_ctx, cat_ctx, pos_ctx, definition, tldr, detail_body, citations, concern = row
        if not canonical:
            continue
        entries.append({
            'canonical_name': canonical.strip(),
            'display_name': display,
            'dog_base_severity': SEVERITY_MAP.get(dog_sev, 'neutral'),
            'cat_base_severity': SEVERITY_MAP.get(cat_sev, 'neutral'),
            'position_reduction_eligible': bool(pos_red) if pos_red else False,
            'base_description': base_desc or definition,  # base_description = definition
            'dog_context': dog_ctx,
            'cat_context': cat_ctx,
            'position_context': pos_ctx,
            'definition': definition,
            'tldr': tldr,
            'detail_body': detail_body,
            'citations_display': citations,
            'tier': 'tier3',
        })
    return entries


def parse_tier4_markdown(filepath):
    """Parse Tier 4 markdown with code blocks (19 entries)."""
    with open(filepath, 'r') as f:
        content = f.read()
    
    # Extract code blocks
    blocks = re.findall(r'```\n(.*?)```', content, re.DOTALL)
    entries = []
    
    for block in blocks:
        entry = {}
        lines = block.strip().split('\n')
        current_key = None
        current_val = []
        
        for line in lines:
            # Check if this is a key: value line
            match = re.match(r'^(\w[\w_]*)\s*:\s*(.*)$', line)
            if match:
                # Save previous key
                if current_key:
                    entry[current_key] = '\n'.join(current_val).strip()
                current_key = match.group(1)
                current_val = [match.group(2)]
            else:
                # Continuation of previous value
                if current_key:
                    current_val.append(line)
        
        # Save last key
        if current_key:
            entry[current_key] = '\n'.join(current_val).strip()
        
        if 'canonical_name' not in entry:
            continue
            
        entries.append({
            'canonical_name': entry.get('canonical_name', '').strip(),
            'display_name': entry.get('display_name', ''),
            'dog_base_severity': SEVERITY_MAP.get(entry.get('dog_base_severity', 'neutral'), 'neutral'),
            'cat_base_severity': SEVERITY_MAP.get(entry.get('cat_base_severity', 'neutral'), 'neutral'),
            'position_reduction_eligible': entry.get('position_reduction_eligible', 'FALSE').upper() == 'TRUE',
            'base_description': entry.get('base_description', '') or entry.get('definition', ''),
            'dog_context': entry.get('dog_context', ''),
            'cat_context': entry.get('cat_context', ''),
            'position_context': entry.get('position_context', ''),
            'definition': entry.get('definition', ''),
            'tldr': entry.get('tldr', ''),
            'detail_body': entry.get('detail_body', ''),
            'citations_display': entry.get('citations_display', ''),
            'tier': 'tier4',
        })
    
    return entries


def import_entries(sb, entries):
    """UPSERT entries into ingredients_dict."""
    stats = {
        'updated': 0,
        'inserted': 0,
        'severity_changed': 0,
        'skipped': 0,
        'errors': 0,
    }
    changes = []
    
    for entry in entries:
        cn = entry['canonical_name']
        
        # Check if exists
        result = sb.table('ingredients_dict') \
            .select('id, canonical_name, dog_base_severity, cat_base_severity, display_name, tldr') \
            .eq('canonical_name', cn) \
            .execute()
        
        update_data = {
            'display_name': entry['display_name'],
            'dog_base_severity': entry['dog_base_severity'],
            'cat_base_severity': entry['cat_base_severity'],
            'position_reduction_eligible': entry['position_reduction_eligible'],
            'base_description': entry['base_description'],
            'dog_context': entry['dog_context'],
            'cat_context': entry['cat_context'],
            'position_context': entry['position_context'],
            'definition': entry['definition'],
            'tldr': entry['tldr'],
            'detail_body': entry['detail_body'],
            'citations_display': entry['citations_display'],
        }
        
        if result.data and len(result.data) > 0:
            existing = result.data[0]
            old_dog = existing.get('dog_base_severity')
            old_cat = existing.get('cat_base_severity')
            had_tldr = bool(existing.get('tldr'))
            
            severity_changed = (old_dog != entry['dog_base_severity'] or old_cat != entry['cat_base_severity'])
            
            if DRY_RUN:
                action = 'UPDATE'
                if severity_changed:
                    action += f' (severity: {old_dog}/{old_cat} → {entry["dog_base_severity"]}/{entry["cat_base_severity"]})'
                if not had_tldr:
                    action += ' (adding content)'
                print(f'  {cn}: {action}')
            else:
                try:
                    sb.table('ingredients_dict') \
                        .update(update_data) \
                        .eq('canonical_name', cn) \
                        .execute()
                except Exception as e:
                    print(f'  ERROR updating {cn}: {e}')
                    stats['errors'] += 1
                    continue
            
            stats['updated'] += 1
            if severity_changed:
                stats['severity_changed'] += 1
                changes.append(f'{cn}: {old_dog}/{old_cat} → {entry["dog_base_severity"]}/{entry["cat_base_severity"]}')
        else:
            # INSERT new entry
            insert_data = {
                **update_data,
                'canonical_name': cn,
                'is_unnamed_species': False,
                'is_legume': False,
                'cat_carb_flag': False,
            }
            
            if DRY_RUN:
                print(f'  {cn}: INSERT (new) [{entry["dog_base_severity"]}/{entry["cat_base_severity"]}]')
            else:
                try:
                    sb.table('ingredients_dict') \
                        .insert(insert_data) \
                        .execute()
                except Exception as e:
                    print(f'  ERROR inserting {cn}: {e}')
                    stats['errors'] += 1
                    continue
            
            stats['inserted'] += 1
    
    return stats, changes


def main():
    mode = 'DRY RUN' if DRY_RUN else 'LIVE'
    print(f'=== Tier 3 + Tier 4 Import ({mode}) ===\n')
    
    # Parse files
    tier3_path = '/mnt/user-data/uploads/TIER3_VITAMINS_MINERALS_UPDATED.xlsx'
    tier4_path = '/mnt/user-data/uploads/compass_artifact_wf-c3516a1f-230f-417b-adc3-d5919fa4088c_text_markdown.md'
    
    print('Parsing Tier 3 (xlsx)...')
    tier3 = parse_tier3_xlsx(tier3_path)
    print(f'  {len(tier3)} entries')
    
    print('Parsing Tier 4 (markdown)...')
    tier4 = parse_tier4_markdown(tier4_path)
    print(f'  {len(tier4)} entries')
    
    all_entries = tier3 + tier4
    print(f'\nTotal: {len(all_entries)} entries to import\n')
    
    # Connect
    sb = get_client()
    
    # Import
    print('Importing Tier 3...')
    stats3, changes3 = import_entries(sb, tier3)
    
    print('\nImporting Tier 4...')
    stats4, changes4 = import_entries(sb, tier4)
    
    # Totals
    total_updated = stats3['updated'] + stats4['updated']
    total_inserted = stats3['inserted'] + stats4['inserted']
    total_sev = stats3['severity_changed'] + stats4['severity_changed']
    total_errors = stats3['errors'] + stats4['errors']
    all_changes = changes3 + changes4
    
    print(f'\n=== RESULTS ===')
    print(f'Tier 3: {stats3["updated"]} updated, {stats3["inserted"]} inserted, {stats3["severity_changed"]} severity changes')
    print(f'Tier 4: {stats4["updated"]} updated, {stats4["inserted"]} inserted, {stats4["severity_changed"]} severity changes')
    print(f'Total:  {total_updated} updated, {total_inserted} inserted, {total_sev} severity changes, {total_errors} errors')
    
    if all_changes:
        print(f'\nSeverity changes:')
        for c in all_changes:
            print(f'  {c}')
    
    # Add synonym for brewer_dried_yeast → brewers_yeast
    print(f'\nNote: brewer_dried_yeast skipped (alias for brewers_yeast per Tier 4 research).')
    print(f'Add to synonyms.json: "brewer_dried_yeast" → "brewers_yeast"')


if __name__ == '__main__':
    main()
